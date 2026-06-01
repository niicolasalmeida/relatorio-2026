#!/usr/bin/env python3
from __future__ import annotations

import calendar
import datetime as dt
import json
from pathlib import Path

try:
    import openpyxl
except ModuleNotFoundError:
    import xlsx_compat as openpyxl

ROOT = Path(__file__).resolve().parent
BASE_XLSX = ROOT / "Base.xlsx"
OUT = ROOT / "data.inline.js"
FOCUS_DATE_OVERRIDE = dt.date(2026, 5, 1)
SPECIAL_SALE_GROUPS = [
    {
        "date": dt.date(2026, 5, 7),
        "nfs": {"18", "19"},
        "clients": {
            "RUA DOS ALPES EMPREENDIMENTOS IMOBILIARIOS (026444)",
            "CBR MAGIK LZ 14 EMPREENDIMENTOS IMOB. (026471)",
        },
        "group_id": "TRINITY::2026-05-07::RUA-DOS-ALPES-CBR-MAGIK",
        "display_client": "Rua dos Alpes / CBR Magik",
        "display_nf": "18 + 19",
    },
]

MONTHS = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
MONTH_NAME_TO_NUMBER = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}
REGION_BY_UF = {
    "AC": "Norte",
    "AL": "Nordeste",
    "AM": "Norte",
    "AP": "Norte",
    "BA": "Nordeste",
    "CE": "Nordeste",
    "DF": "Centro-Oeste",
    "ES": "Sudeste",
    "GO": "Centro-Oeste",
    "MA": "Nordeste",
    "MG": "Sudeste",
    "MS": "Centro-Oeste",
    "MT": "Centro-Oeste",
    "PA": "Norte",
    "PB": "Nordeste",
    "PE": "Nordeste",
    "PI": "Nordeste",
    "PR": "Sul",
    "RJ": "Sudeste",
    "RN": "Nordeste",
    "RO": "Norte",
    "RR": "Norte",
    "RS": "Sul",
    "SC": "Sul",
    "SE": "Nordeste",
    "SP": "Sudeste",
    "TO": "Norte",
}


def clean_text(value: object, default: str = "") -> str:
    return str(value or "").strip() or default


def as_number(value: object) -> float:
    if value in (None, "", "-"):
        return 0.0
    return float(value)


def as_date(value: object) -> dt.date | None:
    if value in (None, "", "-"):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def week_of_month(value: dt.date) -> int:
    if value.year == 2026 and value.month == 5:
        if value.day <= 9:
            return 1
        if value.day <= 16:
            return 2
        if value.day <= 23:
            return 3
        if value.day <= 30:
            return 4
        return 5
    # Calendar weeks inside the month, using Sunday-Saturday boundaries.
    first_day = value.replace(day=1)
    start_offset = (first_day.weekday() + 1) % 7
    return ((value.day + start_offset - 1) // 7) + 1


def week_bounds_in_month(value: dt.date) -> tuple[dt.date, dt.date]:
    if value.year == 2026 and value.month == 5:
        if value.day <= 9:
            return dt.date(2026, 5, 1), dt.date(2026, 5, 9)
        if value.day <= 16:
            return dt.date(2026, 5, 10), dt.date(2026, 5, 16)
        if value.day <= 23:
            return dt.date(2026, 5, 17), dt.date(2026, 5, 23)
        if value.day <= 30:
            return dt.date(2026, 5, 24), dt.date(2026, 5, 30)
        return dt.date(2026, 5, 31), dt.date(2026, 5, 31)
    month_start = value.replace(day=1)
    month_end = value.replace(day=calendar.monthrange(value.year, value.month)[1])
    days_since_sunday = (value.weekday() + 1) % 7
    week_start = value - dt.timedelta(days=days_since_sunday)
    week_end = week_start + dt.timedelta(days=6)
    if week_start < month_start:
        week_start = month_start
    if week_end > month_end:
        week_end = month_end
    return week_start, week_end


def month_meta(value: dt.date) -> dict[str, object]:
    return {
        "Ano": value.year,
        "MesNumero": value.month,
        "MesNome": MONTHS[value.month - 1],
        "AnoMes": f"{value.year:04d}-{value.month:02d}",
        "SemanaMes": week_of_month(value),
    }


def month_number_from_label(value: object) -> int | None:
    if value in (None, "", "-"):
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.month
    text = str(value).strip()
    if not text:
        return None
    lowered = text.lower()
    if lowered.isdigit():
        month = int(lowered)
        return month if 1 <= month <= 12 else None
    return MONTH_NAME_TO_NUMBER.get(lowered)


def focus_params(value: dt.date) -> dict[str, str]:
    week_index = week_of_month(value)
    start_date, end_date = week_bounds_in_month(value)
    return {
        "AnoFoco": str(value.year),
        "MesFoco": str(value.month),
        "SemanaIndice": str(week_index),
        "SemanaInicio": start_date.isoformat(),
        "SemanaFim": end_date.isoformat(),
        "SemanaLabel": f"{start_date.strftime('%d/%m')}–{end_date.strftime('%d/%m')}",
    }


def special_sale_group_for(date_value: dt.date, nf: str, client: str) -> dict[str, object] | None:
    for group in SPECIAL_SALE_GROUPS:
        if date_value != group["date"]:
            continue
        if nf not in group["nfs"]:
            continue
        if client not in group["clients"]:
            continue
        return group
    return None


def build_vendas(wb: openpyxl.Workbook) -> list[dict[str, object]]:
    ws = wb["Base"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = [
        "Nr. nota",
        "Pessoa",
        "Dt. emissão",
        "Tipo",
        "Linha de Receita",
        "Vl. total vendido",
        "Cidade",
        "UF",
        "Estado",
        "Detalhamento_Produto",
        "Classficação Produto",
        "Vendedor",
        "Forma de Pagamento",
        "Origem do LEAD",
    ]
    missing = [name for name in required if name not in idx]
    if missing:
        raise SystemExit(f"Colunas ausentes na aba Base: {', '.join(missing)}")

    rows: list[dict[str, object]] = []
    for sale_id, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        date_value = as_date(row[idx["Dt. emissão"]])
        if date_value is None:
            continue
        meta = month_meta(date_value)
        uf = clean_text(row[idx["UF"]]).upper()
        nf = clean_text(row[idx["Nr. nota"]])
        client = clean_text(row[idx["Pessoa"]], "Nao informado")
        special_group = special_sale_group_for(date_value, nf, client)
        rows.append({
            "VendaID": str(sale_id),
            "DataEmissao": date_value.isoformat(),
            **meta,
            "NF": nf,
            "Tipo": clean_text(row[idx["Tipo"]], "Nao informado"),
            "LinhaReceita": clean_text(row[idx["Linha de Receita"]], "Nao informado"),
            "Valor": round(as_number(row[idx["Vl. total vendido"]]), 2),
            "Cliente": client,
            "Cidade": clean_text(row[idx["Cidade"]], "Nao informado"),
            "UF": uf,
            "Estado": clean_text(row[idx["Estado"]], "Nao informado"),
            "DetalhamentoProduto": clean_text(row[idx["Detalhamento_Produto"]], "Nao informado"),
            "ClassificacaoProduto": clean_text(row[idx["Classficação Produto"]], "Nao informado"),
            "Vendedor": clean_text(row[idx["Vendedor"]]),
            "FormaPagamento": clean_text(row[idx["Forma de Pagamento"]], "Não informado"),
            "OrigemLead": clean_text(row[idx["Origem do LEAD"]], "Não informado"),
            "Regiao": REGION_BY_UF.get(uf, "Nao informado"),
            "VendaAgrupadaId": str(special_group["group_id"]) if special_group else "",
            "VendaAgrupadaCliente": str(special_group["display_client"]) if special_group else "",
            "VendaAgrupadaNF": str(special_group["display_nf"]) if special_group else "",
        })
    return rows


def build_financeiro(wb: openpyxl.Workbook) -> list[dict[str, object]]:
    ws = wb["Financeiro"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = [
        "Nome Cliente",
        "Tipo de Pagamento",
        "Valor da Entrada",
        "Data da Entrada",
        "Quantidade de Parcelas",
        "Valor da Parcela",
        "Data da 1ª Parcela",
    ]
    missing = [name for name in required if name not in idx]
    if missing:
        raise SystemExit(f"Colunas ausentes na aba Financeiro: {', '.join(missing)}")

    rows: list[dict[str, object]] = []
    event_id = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        client = clean_text(row[idx["Nome Cliente"]], "Nao informado")
        payment_type = clean_text(row[idx["Tipo de Pagamento"]], "Nao informado")
        entry_value = round(as_number(row[idx["Valor da Entrada"]]), 2)
        entry_date = as_date(row[idx["Data da Entrada"]])
        parcel_count = int(as_number(row[idx["Quantidade de Parcelas"]]))
        parcel_value = round(as_number(row[idx["Valor da Parcela"]]), 10)
        first_parcel_date = as_date(row[idx["Data da 1ª Parcela"]])

        if entry_date and entry_value > 0:
          meta = month_meta(entry_date)
          rows.append({
              "EventoID": str(event_id),
              "Cliente": client,
              "EventoTipo": "Entrada",
              "DataEvento": entry_date.isoformat(),
              **meta,
              "Valor": entry_value,
              "TipoPagamentoOrigem": payment_type,
              "QuantidadeParcelasOrigem": str(parcel_count),
              "ValorParcelaOrigem": parcel_value,
              "DataEntradaOrigem": entry_date.isoformat(),
              "DataPrimeiraParcelaOrigem": first_parcel_date.isoformat() if first_parcel_date else "",
          })
          event_id += 1

        if first_parcel_date and parcel_count > 0 and parcel_value > 0:
            for parcel_idx in range(parcel_count):
                target_year = first_parcel_date.year + ((first_parcel_date.month - 1 + parcel_idx) // 12)
                target_month = ((first_parcel_date.month - 1 + parcel_idx) % 12) + 1
                target_day = min(
                    first_parcel_date.day,
                    calendar.monthrange(target_year, target_month)[1],
                )
                parcel_date = dt.date(target_year, target_month, target_day)
                meta = month_meta(parcel_date)
                rows.append({
                    "EventoID": str(event_id),
                    "Cliente": client,
                    "EventoTipo": "Parcela",
                    "DataEvento": parcel_date.isoformat(),
                    **meta,
                    "Valor": parcel_value,
                    "TipoPagamentoOrigem": payment_type,
                    "QuantidadeParcelasOrigem": str(parcel_count),
                    "ValorParcelaOrigem": parcel_value,
                    "DataEntradaOrigem": entry_date.isoformat() if entry_date else "",
                    "DataPrimeiraParcelaOrigem": first_parcel_date.isoformat(),
                })
                event_id += 1

    return rows


def build_estoque(wb: openpyxl.Workbook, default_year: int) -> list[dict[str, object]]:
    ws = wb["Estoque"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = [
        "Mês",
        "Empresa",
        "Setor",
        "Grupo de produto",
        "Família de produto",
        "Código",
        "Nome do produto",
        "Quantidade",
        "Qt. futura",
        "Qt. requisitada",
        "Preço de Custo Splan",
        "CUSTO",
    ]
    missing = [name for name in required if name not in idx]
    if missing:
        raise SystemExit(f"Colunas ausentes na aba Estoque: {', '.join(missing)}")

    rows_by_code: dict[tuple[int, int, str], dict[str, object]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = clean_text(row[idx["Código"]])
        if not code:
            continue
        month_number = month_number_from_label(row[idx["Mês"]])
        if month_number is None:
            continue

        quantity = as_number(row[idx["Quantidade"]])
        future_qty = as_number(row[idx["Qt. futura"]])
        requested_qty = as_number(row[idx["Qt. requisitada"]])
        unit_cost = as_number(row[idx["Preço de Custo Splan"]])
        total_cost = as_number(row[idx["CUSTO"]])
        sector = clean_text(row[idx["Setor"]]).upper()

        item = rows_by_code.setdefault((default_year, month_number, code), {
            "Ano": default_year,
            "MesNumero": month_number,
            "MesNome": MONTHS[month_number - 1],
            "AnoMes": f"{default_year:04d}-{month_number:02d}",
            "Empresa": clean_text(row[idx["Empresa"]], "Nao informado"),
            "Codigo": code,
            "Produto": clean_text(row[idx["Nome do produto"]], "Nao informado"),
            "Familia": clean_text(row[idx["Família de produto"]], "Nao informado"),
            "Grupo": clean_text(row[idx["Grupo de produto"]], "Nao informado"),
            "Quantidade": 0.0,
            "EstoquePrincipal": 0.0,
            "EstoqueDemonstracao": 0.0,
            "QtFutura": 0.0,
            "QtRequisitada": 0.0,
            "CustoUnitario": 0.0,
            "CustoTotal": 0.0,
        })

        item["Quantidade"] = round(float(item["Quantidade"]) + quantity, 2)
        item["QtFutura"] = round(float(item["QtFutura"]) + future_qty, 2)
        item["QtRequisitada"] = round(float(item["QtRequisitada"]) + requested_qty, 2)
        item["CustoTotal"] = round(float(item["CustoTotal"]) + total_cost, 2)
        if unit_cost > 0:
            item["CustoUnitario"] = round(unit_cost, 2)

        if "PRINCIPAL" in sector:
            item["EstoquePrincipal"] = round(float(item["EstoquePrincipal"]) + quantity, 2)
        elif "DEMONSTRA" in sector:
            item["EstoqueDemonstracao"] = round(float(item["EstoqueDemonstracao"]) + quantity, 2)

    return [rows_by_code[key] for key in sorted(rows_by_code)]


def main() -> None:
    if not BASE_XLSX.exists():
        raise SystemExit(f"Arquivo nao encontrado: {BASE_XLSX}")

    wb = openpyxl.load_workbook(BASE_XLSX, read_only=True, data_only=True)
    vendas = build_vendas(wb)
    financeiro = build_financeiro(wb)
    if FOCUS_DATE_OVERRIDE is not None:
        focus_date = FOCUS_DATE_OVERRIDE
    elif vendas:
        focus_date = max(dt.date.fromisoformat(row["DataEmissao"]) for row in vendas)
    elif financeiro:
        focus_date = max(dt.date.fromisoformat(row["DataEvento"]) for row in financeiro)
    else:
        focus_date = dt.date.today()

    estoque = build_estoque(wb, focus_date.year)

    payload = {
        "vendas": vendas,
        "financeiro": financeiro,
        "estoque": estoque,
        "parametros": [focus_params(focus_date)],
    }
    OUT.write_text(
        "window.__DASHBOARD_DATA__ = " + json.dumps(payload, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(f"Gerado: {OUT}")


if __name__ == "__main__":
    main()
