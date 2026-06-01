#!/usr/bin/env python3
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

try:
    import openpyxl
except ModuleNotFoundError:
    import xlsx_compat as openpyxl

ROOT = Path(__file__).resolve().parent
BASE_XLSX = ROOT / "Base.xlsx"
OUT = ROOT / "projecao.inline.js"

MONTHS = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def clean_text(value: object, default: str = "") -> str:
    return str(value or "").strip() or default


def as_number(value: object) -> float:
    if value in (None, "", "-"):
        return 0.0
    return float(value)


def as_month_date(value: object) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date().replace(day=1)
    if isinstance(value, dt.date):
        return value.replace(day=1)
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text, fmt).date().replace(day=1)
        except ValueError:
            continue
    return None


def month_meta(value: dt.date) -> dict[str, object]:
    return {
        "Ano": value.year,
        "MesNumero": value.month,
        "MesNome": MONTHS[value.month - 1],
        "AnoMes": f"{value.year:04d}-{value.month:02d}",
    }


def find_row(ws: openpyxl.worksheet.worksheet.Worksheet, label: str) -> int | None:
    target = clean_text(label).casefold()
    for row_idx in range(1, ws.max_row + 1):
        value = clean_text(ws.cell(row_idx, 1).value)
        if value.casefold() == target:
            return row_idx
    return None


def build_receipts_block(ws: openpyxl.worksheet.worksheet.Worksheet, title_row: int) -> list[dict[str, object]]:
    company = clean_text(ws.cell(title_row, 1).value, "Nao informado").title()
    header_row = title_row + 1
    month_cols: list[tuple[int, dt.date]] = []
    for col in range(2, ws.max_column + 1):
        month_date = as_month_date(ws.cell(header_row, col).value)
        if month_date:
            month_cols.append((col, month_date))

    rows: list[dict[str, object]] = []
    row_idx = header_row + 1
    while row_idx <= ws.max_row:
        label = clean_text(ws.cell(row_idx, 1).value)
        if not label:
            row_idx += 1
            continue
        if label.casefold() == "total":
            break

        for col, month_date in month_cols:
            value = round(as_number(ws.cell(row_idx, col).value), 2)
            if value == 0:
                continue
            rows.append({
                "Empresa": company,
                "Linha": label,
                **month_meta(month_date),
                "DataMes": month_date.isoformat(),
                "Valor": value,
            })
        row_idx += 1

    return rows


def build_projection(wb: openpyxl.Workbook) -> dict[str, list[dict[str, object]]]:
    ws = wb["Projeção"]

    receipts = []
    for title in ("MIHA", "SPLAN"):
        title_row = find_row(ws, title)
        if title_row:
            receipts.extend(build_receipts_block(ws, title_row))

    payments: list[dict[str, object]] = []
    month_pairs: list[tuple[int, dt.date]] = []
    payment_title_row = find_row(ws, "Projeçao Pagamentos")
    if payment_title_row is None:
        payment_title_row = find_row(ws, "Projeção Pagamentos")
    if payment_title_row is None:
        return {
            "recebimentos": receipts,
            "pagamentos": payments,
        }

    payment_first_category_row = payment_title_row + 3

    for col in range(2, ws.max_column + 1, 2):
        month_date = as_month_date(ws.cell(payment_title_row, col).value)
        if month_date:
            month_pairs.append((col, month_date))

    for row_idx in range(payment_first_category_row, ws.max_row + 1):
        category = clean_text(ws.cell(row_idx, 1).value)
        if not category:
            continue
        for budget_col, month_date in month_pairs:
            budget_value = round(abs(as_number(ws.cell(row_idx, budget_col).value)), 2)
            realized_value = round(abs(as_number(ws.cell(row_idx, budget_col + 1).value)), 2)
            if budget_value == 0 and realized_value == 0:
                continue
            payments.append({
                "Categoria": category,
                **month_meta(month_date),
                "DataMes": month_date.isoformat(),
                "ValorOrcado": budget_value,
                "ValorRealizado": realized_value,
            })

    return {
        "recebimentos": receipts,
        "pagamentos": payments,
    }


def main() -> None:
    if not BASE_XLSX.exists():
        raise SystemExit(f"Arquivo nao encontrado: {BASE_XLSX}")

    wb = openpyxl.load_workbook(BASE_XLSX, read_only=True, data_only=True)
    payload = build_projection(wb)

    js = (
        "(function () {\n"
        "  const data = window.__DASHBOARD_DATA__ || {};\n"
        f"  data.projecao = {json.dumps(payload, ensure_ascii=False)};\n"
        "  window.__DASHBOARD_DATA__ = data;\n"
        "})();\n"
    )
    OUT.write_text(js, encoding="utf-8")
    print(f"Gerado: {OUT}")


if __name__ == "__main__":
    main()
