#!/usr/bin/env python3
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
BASE_XLSX = ROOT / "Base.xlsx"
OUT = ROOT / "despesas.inline.js"

MONTHS = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def as_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def as_number(value) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def week_of_month(value: dt.date) -> int:
    return ((value.day - 1) // 7) + 1


def main() -> None:
    if not BASE_XLSX.exists():
      raise SystemExit(f"Arquivo nao encontrado: {BASE_XLSX}")

    wb = openpyxl.load_workbook(BASE_XLSX, read_only=True, data_only=True)
    ws = wb["Despesas"]

    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = [
        "Cód. empresa",
        "Nome da pessoa",
        "Forma de pagamento",
        "Dt. emissão",
        "Dt. baixa",
        "Vl. líquido",
        "Vl. em aberto",
        "Classificação 0",
        "Classificação 1",
        "Classificação 2",
        "Caixa",
    ]
    missing = [name for name in required if name not in idx]
    if missing:
        raise SystemExit(f"Colunas ausentes na aba Despesas: {', '.join(missing)}")

    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        issue_date = as_date(row[idx["Dt. emissão"]])
        if issue_date is None:
            continue
        baixa_date = as_date(row[idx["Dt. baixa"]])

        year = issue_date.year
        month = issue_date.month
        rows.append({
            "Empresa": str(row[idx["Cód. empresa"]] or "").strip(),
            "Favorecido": str(row[idx["Nome da pessoa"]] or "").strip(),
            "Classificacao0": str(row[idx["Classificação 0"]] or "").strip(),
            "Classificacao1": str(row[idx["Classificação 1"]] or "").strip(),
            "Natureza": str(row[idx["Classificação 2"]] or "").strip(),
            "Classificacao2": str(row[idx["Classificação 2"]] or "").strip(),
            "Caixa": str(row[idx["Caixa"]] or "").strip(),
            "ClassificacaoCaixa": str(row[idx["Caixa"]] or "").strip(),
            "FormaPagamento": str(row[idx["Forma de pagamento"]] or "").strip(),
            "Ano": year,
            "MesNumero": month,
            "MesNome": MONTHS[month - 1],
            "AnoMes": f"{year:04d}-{month:02d}",
            "SemanaMes": week_of_month(issue_date),
            "DataEmissao": issue_date.isoformat(),
            "AnoBaixa": baixa_date.year if baixa_date else 0,
            "MesNumeroBaixa": baixa_date.month if baixa_date else 0,
            "MesNomeBaixa": MONTHS[baixa_date.month - 1] if baixa_date else "",
            "AnoMesBaixa": f"{baixa_date.year:04d}-{baixa_date.month:02d}" if baixa_date else "",
            "SemanaMesBaixa": week_of_month(baixa_date) if baixa_date else 0,
            "DataBaixa": baixa_date.isoformat() if baixa_date else "",
            "ValorLiquido": abs(as_number(row[idx["Vl. líquido"]])),
            "ValorAberto": abs(as_number(row[idx["Vl. em aberto"]])),
        })

    js = (
        "(function () {\n"
        "  const data = window.__DASHBOARD_DATA__ || {};\n"
        f"  data.despesas = {json.dumps(rows, ensure_ascii=False)};\n"
        "  window.__DASHBOARD_DATA__ = data;\n"
        "})();\n"
    )
    OUT.write_text(js, encoding="utf-8")
    print(f"Gerado: {OUT}")


if __name__ == "__main__":
    main()
