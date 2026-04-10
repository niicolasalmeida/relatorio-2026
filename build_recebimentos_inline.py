#!/usr/bin/env python3
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
BASE_XLSX = ROOT / "Base.xlsx"
OUT = ROOT / "recebimentos.inline.js"

MONTHS = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def as_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def as_number(value) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def week_of_month(value: dt.date) -> int:
    first_day = value.replace(day=1)
    start_offset = (first_day.weekday() + 1) % 7
    return ((value.day + start_offset - 1) // 7) + 1


def main() -> None:
    if not BASE_XLSX.exists():
        raise SystemExit(f"Arquivo nao encontrado: {BASE_XLSX}")

    wb = openpyxl.load_workbook(BASE_XLSX, read_only=True, data_only=True)
    ws = wb["Recebimento"]

    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = [
        "Nome da pessoa",
        "Empresa",
        "Dt. baixa",
        "Vl. líquido",
        "Vl. baixado",
        "Classificação 1",
        "Classificação 2",
        "Classificação Caixa",
        "Natureza de lançamento",
    ]
    missing = [name for name in required if name not in idx]
    if missing:
        raise SystemExit(f"Colunas ausentes na aba Recebimento: {', '.join(missing)}")

    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_value = as_date(row[idx["Dt. baixa"]])
        if date_value is None:
            continue

        year = date_value.year
        month = date_value.month
        rows.append({
            "Cliente": str(row[idx["Nome da pessoa"]] or "").strip(),
            "Empresa": str(row[idx["Empresa"]] or "").strip(),
            "Natureza": str(row[idx["Natureza de lançamento"]] or "").strip(),
            "Classificacao1": str(row[idx["Classificação 1"]] or "").strip(),
            "Classificacao2": str(row[idx["Classificação 2"]] or "").strip(),
            "ClassificacaoCaixa": str(row[idx["Classificação Caixa"]] or "").strip(),
            "Ano": year,
            "MesNumero": month,
            "MesNome": MONTHS[month - 1],
            "AnoMes": f"{year:04d}-{month:02d}",
            "SemanaMes": week_of_month(date_value),
            "DataBaixa": date_value.isoformat(),
            "ValorLiquido": abs(as_number(row[idx["Vl. líquido"]])),
            "ValorBaixado": abs(as_number(row[idx["Vl. baixado"]])),
        })

    js = (
        "(function () {\n"
        "  const data = window.__DASHBOARD_DATA__ || {};\n"
        f"  data.recebimentos = {json.dumps(rows, ensure_ascii=False)};\n"
        "  window.__DASHBOARD_DATA__ = data;\n"
        "})();\n"
    )
    OUT.write_text(js, encoding="utf-8")
    print(f"Gerado: {OUT}")


if __name__ == "__main__":
    main()
